from pathlib import Path
from typing import Type, Union

from openpyxl.reader.excel import ExcelReader

from .schemas import SheetTemplate

KEEP_VBA = False


def load_workbook(
        filename, read_only=False, keep_vba=KEEP_VBA,
        data_only=False, keep_links=True, excel_reader=ExcelReader
):
    reader = excel_reader(filename, read_only, keep_vba,
                          data_only, keep_links)
    reader.read()
    return reader.wb


def validate_xlsx(
        fp: Union[str, Path], model: Type[SheetTemplate],
        sheet_index: int = 0, return_validate_errors=False,
        excel_reader=ExcelReader, **kwargs
):
    wb = load_workbook(fp, excel_reader=excel_reader)
    sheet = wb.worksheets[sheet_index]
    rows = sheet.iter_rows()
    headers = {th.value: i for i, th in enumerate(next(rows))}

    for row_x, raw_row in enumerate(rows, start=2):
        try:
            item = model.validate({key: v for key, v in zip(headers, raw_row) if key is not None},)
        except BaseException as e:
            if return_validate_errors:
                yield e
            raise e

        yield item


if __name__ == 'main':
    pass
